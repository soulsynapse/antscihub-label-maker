#!/usr/bin/env python
"""
PyQt6 GUI for building Avery 5267 label sheets.

Layout:
- Tab 1: Import CSV/Excel data with a left-side table preview
- Tab 2: Label builder with live sheet preview and settings
"""

from __future__ import absolute_import, division, print_function

import csv
import json
import os
import re
import sys
import tempfile
import traceback
from datetime import date, datetime

from PyQt6.QtCore import QRectF, QSettings, Qt, pyqtSignal
from PyQt6.QtGui import (
    QAction,
    QColor,
    QFont,
    QFontMetrics,
    QImage,
    QImageReader,
    QPageLayout,
    QPageSize,
    QPainter,
    QPen,
)
from PyQt6.QtPrintSupport import QPrintDialog, QPrinter
from PyQt6.QtWidgets import (
    QAbstractItemView,
    QApplication,
    QCheckBox,
    QColorDialog,
    QComboBox,
    QDialog,
    QFileDialog,
    QFormLayout,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QSpinBox,
    QSplitter,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

try:
    import cv2
except ImportError:
    cv2 = None

try:
    from pylibdmtx.pylibdmtx import encode as dmtx_encode
except Exception:  # pylint: disable=broad-except
    dmtx_encode = None

try:
    import qrcode
except Exception:  # pylint: disable=broad-except
    qrcode = None

try:
    from openpyxl import load_workbook
except Exception:  # pylint: disable=broad-except
    load_workbook = None

# Ensure importing sibling script works when launching from repo root.
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPT_DIR not in sys.path:
    sys.path.insert(0, SCRIPT_DIR)

from generate_avery_5267_labels import (  # noqa: E402
    ARUCO_LEFT_OFFSET_IN,
    ARUCO_MARKER_SIZE_IN,
    ARUCO_RIGHT_GAP_IN,
    COLS,
    DEFAULT_ARUCO_DICT,
    DEFAULT_DPI,
    FONT_POINT_SIZE,
    H_PITCH_IN,
    LABEL_HEIGHT_IN,
    LABEL_WIDTH_IN,
    LEFT_MARGIN_IN,
    MARKER_TYPE_ARUCO,
    MARKER_TYPE_DATAMATRIX,
    MARKER_TYPE_NONE,
    MARKER_TYPE_QR,
    MAX_TEXT_LINES,
    MAX_LABELS,
    PAGE_HEIGHT_IN,
    PAGE_WIDTH_IN,
    ROWS,
    SIDE_LINE_HEIGHT_IN,
    SIDE_LINE_TEXT_GAP_IN,
    SIDE_LINE_WIDTH_IN,
    TEXT_VERTICAL_PADDING_IN,
    TOP_MARGIN_IN,
    V_PITCH_IN,
    generate_avery_5267_sheet,
    generate_avery_5267_text_sheet,
)


ARUCO_DICTS = [
    "DICT_4X4_50",
    "DICT_4X4_100",
    "DICT_4X4_250",
    "DICT_5X5_50",
    "DICT_5X5_100",
    "DICT_5X5_250",
    "DICT_6X6_50",
    "DICT_6X6_100",
    "DICT_6X6_250",
    "DICT_7X7_50",
    "DICT_7X7_100",
    "DICT_7X7_250",
]

MONTH_ABBR_UPPER = {
    1: "JAN",
    2: "FEB",
    3: "MAR",
    4: "APR",
    5: "MAY",
    6: "JUN",
    7: "JUL",
    8: "AUG",
    9: "SEP",
    10: "OCT",
    11: "NOV",
    12: "DEC",
}
TOKEN_PATTERN = re.compile(r"\{(date|name|dict|id)\}", flags=re.IGNORECASE)
MARKER_TYPE_ITEMS = (
    ("ArUco", MARKER_TYPE_ARUCO),
    ("DataMatrix", MARKER_TYPE_DATAMATRIX),
    ("QR", MARKER_TYPE_QR),
)
MARKER_TYPE_VALUE_BY_LABEL = {
    label.strip().lower(): value for label, value in MARKER_TYPE_ITEMS
}
MARKER_TYPE_VALUE_BY_LABEL["none"] = MARKER_TYPE_NONE


def _safe_int(text, fallback):
    try:
        return int(text)
    except (TypeError, ValueError):
        return fallback


def _safe_bool(value, fallback=False):
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in ("1", "true", "yes", "on")
    return fallback


def _safe_float(value, fallback):
    try:
        return float(value)
    except (TypeError, ValueError):
        return fallback


def _marker_value_text(value, fallback=""):
    if value is None:
        return str(fallback)
    text = str(value).strip()
    if text == "":
        return str(fallback)
    return text


def _numeric_marker_id_or_zero(value):
    text = _marker_value_text(value, "")
    if text.isdigit():
        return int(text)
    return 0


def _crop_qimage_to_dark_pixels(qimg):
    min_x = qimg.width()
    min_y = qimg.height()
    max_x = -1
    max_y = -1
    for y in range(qimg.height()):
        for x in range(qimg.width()):
            if qimg.pixelColor(x, y).value() >= 245:
                continue
            min_x = min(min_x, x)
            min_y = min(min_y, y)
            max_x = max(max_x, x)
            max_y = max(max_y, y)

    if max_x < min_x or max_y < min_y:
        return qimg
    return qimg.copy(min_x, min_y, max_x - min_x + 1, max_y - min_y + 1)


def _normalize_marker_type(value):
    if value is None:
        return MARKER_TYPE_NONE
    key = str(value).strip().lower()
    if key in (MARKER_TYPE_ARUCO, MARKER_TYPE_DATAMATRIX, MARKER_TYPE_QR, MARKER_TYPE_NONE):
        return key
    return MARKER_TYPE_VALUE_BY_LABEL.get(key, MARKER_TYPE_NONE)


def _format_today_line5():
    today = date.today()
    month_code = MONTH_ABBR_UPPER.get(today.month, "{0:02d}".format(today.month))
    return "{0:02d}.{1}.{2:04d}".format(today.day, month_code, today.year)


class SheetPreviewWidget(QWidget):
    wheel_adjust_requested = pyqtSignal(int, bool, bool)

    def __init__(self, parent=None):
        super(SheetPreviewWidget, self).__init__(parent)
        self.setMinimumSize(520, 700)
        self.missing = 0
        self.row_specs = []
        self._aruco_cache = {}
        self.zoom_factor = 1.0
        self.pan_x = 0.0
        self.pan_y = 0.0
        self._is_panning = False
        self._last_pan_pos = None

    def mousePressEvent(self, event):  # pylint: disable=invalid-name
        if event.button() == Qt.MouseButton.MiddleButton:
            self._is_panning = True
            self._last_pan_pos = event.position()
            self.setCursor(Qt.CursorShape.ClosedHandCursor)
            event.accept()
            return
        super(SheetPreviewWidget, self).mousePressEvent(event)

    def mouseMoveEvent(self, event):  # pylint: disable=invalid-name
        if self._is_panning and self._last_pan_pos is not None:
            pos = event.position()
            delta = pos - self._last_pan_pos
            self.pan_x += float(delta.x())
            self.pan_y += float(delta.y())
            self._last_pan_pos = pos
            self.update()
            event.accept()
            return
        super(SheetPreviewWidget, self).mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):  # pylint: disable=invalid-name
        if event.button() == Qt.MouseButton.MiddleButton and self._is_panning:
            self._is_panning = False
            self._last_pan_pos = None
            self.unsetCursor()
            event.accept()
            return
        super(SheetPreviewWidget, self).mouseReleaseEvent(event)

    def wheelEvent(self, event):  # pylint: disable=invalid-name
        delta = event.angleDelta().y()
        if delta == 0:
            event.ignore()
            return

        steps = int(delta / 120)
        if steps == 0:
            steps = 1 if delta > 0 else -1

        modifiers = event.modifiers()
        ctrl_held = bool(modifiers & Qt.KeyboardModifier.ControlModifier)
        shift_held = bool(modifiers & Qt.KeyboardModifier.ShiftModifier)
        self.wheel_adjust_requested.emit(steps, ctrl_held, shift_held)
        event.accept()

    def set_preview_data(self, missing, row_specs):
        self.missing = int(missing)
        self.row_specs = row_specs[:]
        self.update()

    def set_zoom_factor(self, value):
        zoom = max(0.35, min(4.0, float(value)))
        if abs(zoom - self.zoom_factor) < 1e-6:
            return
        self.zoom_factor = zoom
        self.update()

    def zoom_by_steps(self, steps):
        if steps == 0:
            return
        self.set_zoom_factor(self.zoom_factor * (1.1 ** int(steps)))

    def reset_zoom(self):
        self.set_zoom_factor(1.0)
        self.pan_x = 0.0
        self.pan_y = 0.0
        self.update()

    def _fit_page_rect(self):
        margin = 18
        area = self.rect().adjusted(margin, margin, -margin, -margin)
        page_ratio = PAGE_WIDTH_IN / PAGE_HEIGHT_IN
        area_ratio = area.width() / float(max(1, area.height()))

        if area_ratio > page_ratio:
            base_h = area.height()
            base_w = int(round(base_h * page_ratio))
        else:
            base_w = area.width()
            base_h = int(round(base_w / page_ratio))

        page_w = int(round(base_w * self.zoom_factor))
        page_h = int(round(base_h * self.zoom_factor))

        max_pan_x = max(0.0, (page_w - area.width()) / 2.0)
        max_pan_y = max(0.0, (page_h - area.height()) / 2.0)
        self.pan_x = max(-max_pan_x, min(max_pan_x, self.pan_x))
        self.pan_y = max(-max_pan_y, min(max_pan_y, self.pan_y))

        x = area.x() + (area.width() - page_w) / 2.0 + self.pan_x
        y = area.y() + (area.height() - page_h) / 2.0 + self.pan_y
        return QRectF(x, y, page_w, page_h)

    def _label_rect(self, slot_index, page_rect):
        row = slot_index // COLS
        col = slot_index % COLS

        x_in = LEFT_MARGIN_IN + (col * H_PITCH_IN)
        y_in = TOP_MARGIN_IN + (row * V_PITCH_IN)

        x = page_rect.x() + (x_in / PAGE_WIDTH_IN) * page_rect.width()
        y = page_rect.y() + (y_in / PAGE_HEIGHT_IN) * page_rect.height()
        w = (LABEL_WIDTH_IN / PAGE_WIDTH_IN) * page_rect.width()
        h = (LABEL_HEIGHT_IN / PAGE_HEIGHT_IN) * page_rect.height()
        return QRectF(x, y, w, h)

    def _get_aruco_qimage(self, side_px, marker_id, dict_name):
        if side_px < 8:
            return None

        marker_id = _numeric_marker_id_or_zero(marker_id)
        key = ("aruco", str(dict_name), int(marker_id), int(side_px))
        cached = self._aruco_cache.get(key)
        if cached is not None:
            return cached

        if cv2 is None or not hasattr(cv2, "aruco"):
            return None
        if not hasattr(cv2.aruco, str(dict_name)):
            return None

        try:
            dictionary_id = getattr(cv2.aruco, str(dict_name))
            dictionary = cv2.aruco.getPredefinedDictionary(dictionary_id)

            # Keep marker IDs valid for the selected dictionary.
            dictionary_size = 0
            if hasattr(dictionary, "bytesList"):
                try:
                    dictionary_size = int(dictionary.bytesList.shape[0])
                except Exception:  # pylint: disable=broad-except
                    dictionary_size = 0
            if dictionary_size > 0:
                marker_id = marker_id % dictionary_size

            if hasattr(cv2.aruco, "generateImageMarker"):
                marker = cv2.aruco.generateImageMarker(dictionary, marker_id, int(side_px))
            elif hasattr(cv2.aruco, "drawMarker"):
                marker = cv2.aruco.drawMarker(dictionary, marker_id, int(side_px))
            else:
                return None

            h, w = marker.shape[:2]
            bytes_per_line = marker.strides[0]
            qimg = QImage(
                marker.data,
                w,
                h,
                bytes_per_line,
                QImage.Format.Format_Grayscale8,
            ).copy()
        except Exception:  # pylint: disable=broad-except
            return None

        self._aruco_cache[key] = qimg
        if len(self._aruco_cache) > 512:
            self._aruco_cache.pop(next(iter(self._aruco_cache)))
        return qimg

    def _get_datamatrix_qimage(self, side_px, marker_id):
        if side_px < 8:
            return None

        marker_value = _marker_value_text(marker_id, "0")
        key = ("datamatrix", marker_value, int(side_px))
        cached = self._aruco_cache.get(key)
        if cached is not None:
            return cached

        if dmtx_encode is None:
            return None

        try:
            encoded = dmtx_encode(marker_value.encode("utf-8"))
            qimg = QImage(
                encoded.pixels,
                int(encoded.width),
                int(encoded.height),
                int(encoded.width) * 3,
                QImage.Format.Format_RGB888,
            ).copy()
            qimg = qimg.convertToFormat(QImage.Format.Format_Grayscale8)
            qimg = _crop_qimage_to_dark_pixels(qimg)
            qimg = qimg.scaled(
                int(side_px),
                int(side_px),
                Qt.AspectRatioMode.IgnoreAspectRatio,
                Qt.TransformationMode.FastTransformation,
            )
        except Exception:  # pylint: disable=broad-except
            return None

        self._aruco_cache[key] = qimg
        if len(self._aruco_cache) > 512:
            self._aruco_cache.pop(next(iter(self._aruco_cache)))
        return qimg

    def _get_qr_qimage(self, side_px, marker_id):
        if side_px < 8:
            return None

        marker_value = _marker_value_text(marker_id, "0")
        key = ("qr", marker_value, int(side_px))
        cached = self._aruco_cache.get(key)
        if cached is not None:
            return cached

        if qrcode is None:
            return None

        try:
            qr = qrcode.QRCode(
                version=None,
                error_correction=qrcode.constants.ERROR_CORRECT_M,
                box_size=1,
                border=0,
            )
            qr.add_data(marker_value)
            qr.make(fit=True)
            matrix = qr.get_matrix()
            if not matrix:
                return None

            module_count = len(matrix)
            qimg = QImage(int(side_px), int(side_px), QImage.Format.Format_Grayscale8)
            qimg.fill(255)
            qp = QPainter(qimg)
            qp.setPen(Qt.PenStyle.NoPen)
            qp.setBrush(QColor(0, 0, 0))
            for y, row_data in enumerate(matrix):
                y0 = int(round((y * side_px) / float(module_count)))
                y1 = int(round(((y + 1) * side_px) / float(module_count)))
                h = max(1, y1 - y0)
                for x, cell_on in enumerate(row_data):
                    if not cell_on:
                        continue
                    x0 = int(round((x * side_px) / float(module_count)))
                    x1 = int(round(((x + 1) * side_px) / float(module_count)))
                    w = max(1, x1 - x0)
                    qp.fillRect(x0, y0, w, h, QColor(0, 0, 0))
            qp.end()
        except Exception:  # pylint: disable=broad-except
            return None

        self._aruco_cache[key] = qimg
        if len(self._aruco_cache) > 512:
            self._aruco_cache.pop(next(iter(self._aruco_cache)))
        return qimg

    def paintEvent(self, event):  # pylint: disable=unused-argument
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.fillRect(self.rect(), QColor(245, 246, 248))

        page_rect = self._fit_page_rect()
        px_per_in = page_rect.width() / float(PAGE_WIDTH_IN)

        # Soft shadow + page.
        shadow_rect = page_rect.adjusted(6, 6, 6, 6)
        painter.fillRect(shadow_rect, QColor(220, 223, 228))
        painter.fillRect(page_rect, QColor(255, 255, 255))
        painter.setPen(QPen(QColor(170, 174, 181), 1))
        painter.drawRect(page_rect)

        # Draw all label outlines.
        painter.setPen(QPen(QColor(228, 230, 234), 1))
        for slot in range(MAX_LABELS):
            painter.drawRect(self._label_rect(slot, page_rect))

        # Used/missing labels shading.
        for slot in range(max(0, min(self.missing, MAX_LABELS))):
            r = self._label_rect(slot, page_rect)
            painter.fillRect(r, QColor(238, 240, 243))
            painter.setPen(QPen(QColor(210, 212, 216), 1))
            painter.drawLine(r.topLeft(), r.bottomRight())
            painter.drawLine(r.topRight(), r.bottomLeft())

        # Active labels and text.
        for seq_index, spec in enumerate(self.row_specs):
            slot = self.missing + seq_index
            if slot >= MAX_LABELS:
                break
            r = self._label_rect(slot, page_rect)
            painter.setPen(QPen(QColor(86, 152, 210), 1.4))
            painter.drawRect(r)

            horizontal_pad = max(1.0, 0.02 * px_per_in)
            text_vertical_pad = TEXT_VERTICAL_PADDING_IN * px_per_in
            font_px = max(1, int(round((FONT_POINT_SIZE / 72.0) * px_per_in)))
            text_x = r.x() + horizontal_pad
            text_right = r.right() - horizontal_pad
            text_w = max(4.0, text_right - text_x)
            text_y0 = r.y() + text_vertical_pad
            text_y1 = (r.y() + r.height()) - text_vertical_pad
            marker_type = _normalize_marker_type(spec.get("marker_type", MARKER_TYPE_NONE))
            marker_value = spec.get("marker_id", spec.get("aruco_id", 0))
            marker_payload = spec.get("marker_payload", spec.get("id_text", marker_value))
            side_line_text = str(spec.get("side_line", "")).strip()

            right_marker_size = ARUCO_MARKER_SIZE_IN * px_per_in
            right_marker_size = min(right_marker_size, r.height())
            right_marker_x = (
                r.x()
                + r.width()
                - (ARUCO_LEFT_OFFSET_IN * px_per_in)
                - right_marker_size
            )
            right_marker_x = max(r.x(), min(r.x() + r.width() - right_marker_size, right_marker_x))
            right_marker_y = r.y() + (r.height() - right_marker_size) / 2.0
            right_marker = QRectF(
                right_marker_x,
                right_marker_y,
                right_marker_size,
                right_marker_size,
            )
            right_marker_side_px = max(16, int(round(right_marker_size)))
            right_marker_img = self._get_datamatrix_qimage(
                right_marker_side_px,
                marker_payload,
            )
            if right_marker_img is not None:
                painter.drawImage(right_marker, right_marker_img)
            else:
                painter.fillRect(right_marker, QColor(18, 18, 18))

            if side_line_text:
                side_w = min(r.width(), SIDE_LINE_WIDTH_IN * px_per_in)
                side_h = min(r.height(), SIDE_LINE_HEIGHT_IN * px_per_in)
                side_x = right_marker.x() - side_w
                side_x = max(r.x(), min(r.x() + r.width() - side_w, side_x))
                side_y = r.y() + (r.height() - side_h) / 2.0
                side_rect = QRectF(side_x, side_y, side_w, side_h)

                painter.fillRect(side_rect, QColor(255, 255, 255))
                side_gap = SIDE_LINE_TEXT_GAP_IN * px_per_in
                text_right = min(text_right, side_rect.x() - side_gap)
                text_w = max(4.0, text_right - text_x)

                painter.save()
                side_font = QFont("Arial")
                side_font.setBold(False)
                side_font.setPixelSize(font_px)
                painter.setFont(side_font)
                painter.setPen(QPen(QColor(0, 0, 0), 1))
                center = side_rect.center()
                painter.translate(center.x(), center.y())
                painter.rotate(-90.0)
                rotated_rect = QRectF(
                    -side_rect.height() / 2.0,
                    -side_rect.width() / 2.0,
                    side_rect.height(),
                    side_rect.width(),
                )
                painter.drawText(
                    rotated_rect,
                    int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter),
                    side_line_text,
                )
                painter.restore()
            else:
                marker_text_gap = ARUCO_RIGHT_GAP_IN * px_per_in
                text_right = min(text_right, right_marker.x() - marker_text_gap)
                text_w = max(4.0, text_right - text_x)

            if marker_type != MARKER_TYPE_NONE:
                marker_size = ARUCO_MARKER_SIZE_IN * px_per_in
                marker_size = min(marker_size, r.height())
                marker_x = r.x() + (ARUCO_LEFT_OFFSET_IN * px_per_in)
                marker_y = r.y() + (r.height() - marker_size) / 2.0
                marker = QRectF(marker_x, marker_y, marker_size, marker_size)
                marker_side_px = max(16, int(round(marker_size)))
                if marker_type == MARKER_TYPE_DATAMATRIX:
                    marker_img = self._get_datamatrix_qimage(marker_side_px, marker_value)
                elif marker_type == MARKER_TYPE_QR:
                    marker_img = self._get_qr_qimage(marker_side_px, marker_value)
                else:
                    marker_img = self._get_aruco_qimage(
                        marker_side_px,
                        _numeric_marker_id_or_zero(marker_value),
                        spec.get("marker_dict", spec.get("aruco_dict", DEFAULT_ARUCO_DICT)),
                    )
                if marker_img is not None:
                    painter.drawImage(marker, marker_img)
                else:
                    painter.fillRect(marker, QColor(18, 18, 18))
                marker_text_gap = ARUCO_RIGHT_GAP_IN * px_per_in
                text_x = marker.x() + marker.width() + marker_text_gap
                text_w = max(4.0, text_right - text_x)

            raw_lines = spec.get("lines", [])
            if isinstance(raw_lines, str):
                raw_lines = [raw_lines]
            lines = [str(line) for line in raw_lines[:MAX_TEXT_LINES]]
            while len(lines) < MAX_TEXT_LINES:
                lines.append("")

            font = QFont("Arial")
            font.setBold(False)
            font.setPixelSize(font_px)
            painter.setFont(font)
            painter.setPen(QPen(QColor(14, 34, 48), 1))
            text_box_h = max(1.0, text_y1 - text_y0)
            line_h = text_box_h / float(MAX_TEXT_LINES)
            for i, line in enumerate(lines[:MAX_TEXT_LINES]):
                if not line.strip():
                    continue
                line_rect = QRectF(
                    text_x,
                    text_y0 + i * line_h,
                    text_w,
                    line_h,
                )
                painter.drawText(
                    line_rect,
                    int(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft),
                    line,
                )

        # Title.
        title_rect = QRectF(page_rect.x(), page_rect.y() - 18, page_rect.width(), 14)
        painter.setPen(QPen(QColor(90, 96, 104), 1))
        painter.setFont(QFont("Arial", 8))
        painter.drawText(
            title_rect,
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            "Avery 5267 preview",
        )

        painter.end()


class TextLabelPreviewWidget(SheetPreviewWidget):
    def __init__(self, parent=None):
        super(TextLabelPreviewWidget, self).__init__(parent)
        self.text_style = {
            "font_size_pt": 14,
            "text_color": "#000000",
            "align": "center",
            "bold": False,
            "italic": False,
        }

    def set_text_preview_data(self, missing, row_specs, style):
        self.missing = int(missing)
        self.row_specs = row_specs[:]
        self.text_style = dict(style)
        self.update()

    def _alignment_flags(self):
        align = str(self.text_style.get("align", "center")).lower()
        if align == "left":
            horizontal = Qt.AlignmentFlag.AlignLeft
        elif align == "right":
            horizontal = Qt.AlignmentFlag.AlignRight
        else:
            horizontal = Qt.AlignmentFlag.AlignHCenter
        return int(horizontal | Qt.AlignmentFlag.AlignVCenter)

    def _text_from_spec(self, spec):
        if not isinstance(spec, dict):
            return str(spec).strip()
        if "text" in spec:
            return str(spec.get("text", "")).strip()
        raw_lines = spec.get("lines", [])
        if isinstance(raw_lines, str):
            return raw_lines.strip()
        if raw_lines is None:
            return ""
        return "\n".join(str(line).strip() for line in raw_lines if str(line).strip())

    def _break_long_word(self, metrics, word, max_width):
        parts = []
        current = ""
        for char in word:
            trial = current + char
            if current and metrics.horizontalAdvance(trial) > max_width:
                parts.append(current)
                current = char
            else:
                current = trial
        if current:
            parts.append(current)
        return parts or [word]

    def _wrap_text(self, metrics, text, max_width):
        wrapped = []
        for paragraph in (text.splitlines() or [text]):
            words = paragraph.split()
            if not words:
                wrapped.append("")
                continue

            current = ""
            for word in words:
                word_parts = (
                    self._break_long_word(metrics, word, max_width)
                    if metrics.horizontalAdvance(word) > max_width
                    else [word]
                )
                for part in word_parts:
                    trial = part if not current else "{0} {1}".format(current, part)
                    if not current or metrics.horizontalAdvance(trial) <= max_width:
                        current = trial
                    else:
                        wrapped.append(current)
                        current = part
            if current:
                wrapped.append(current)
        return wrapped

    def _fit_text_layout(self, text, max_width, max_height, max_font_px):
        base_font = QFont("Arial")
        base_font.setBold(_safe_bool(self.text_style.get("bold", False)))
        base_font.setItalic(_safe_bool(self.text_style.get("italic", False)))
        max_width = max(1, int(round(max_width)))
        max_height = max(1, int(round(max_height)))

        for font_px in range(max(1, int(max_font_px)), 0, -1):
            font = QFont(base_font)
            font.setPixelSize(font_px)
            metrics = QFontMetrics(font)
            lines = self._wrap_text(metrics, text, max_width)
            if not lines:
                return font, [], 0, 0
            line_h = max(1, metrics.height())
            spacing = max(0, int(round(font_px * 0.18))) if len(lines) > 1 else 0
            total_h = (len(lines) * line_h) + ((len(lines) - 1) * spacing)
            max_line_w = max((metrics.horizontalAdvance(line) for line in lines), default=0)
            if total_h <= max_height and max_line_w <= max_width:
                return font, lines, spacing, total_h

        font = QFont(base_font)
        font.setPixelSize(1)
        metrics = QFontMetrics(font)
        lines = self._wrap_text(metrics, text, max_width)
        total_h = len(lines) * max(1, metrics.height())
        return font, lines, 0, total_h

    def paintEvent(self, event):  # pylint: disable=unused-argument
        painter = QPainter(self)
        painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
        painter.fillRect(self.rect(), QColor(245, 246, 248))

        page_rect = self._fit_page_rect()
        px_per_in = page_rect.width() / float(PAGE_WIDTH_IN)

        shadow_rect = page_rect.adjusted(6, 6, 6, 6)
        painter.fillRect(shadow_rect, QColor(220, 223, 228))
        painter.fillRect(page_rect, QColor(255, 255, 255))
        painter.setPen(QPen(QColor(170, 174, 181), 1))
        painter.drawRect(page_rect)

        painter.setPen(QPen(QColor(228, 230, 234), 1))
        for slot in range(MAX_LABELS):
            painter.drawRect(self._label_rect(slot, page_rect))

        for slot in range(max(0, min(self.missing, MAX_LABELS))):
            r = self._label_rect(slot, page_rect)
            painter.fillRect(r, QColor(238, 240, 243))
            painter.setPen(QPen(QColor(210, 212, 216), 1))
            painter.drawLine(r.topLeft(), r.bottomRight())
            painter.drawLine(r.topRight(), r.bottomLeft())

        font_size_pt = _safe_float(self.text_style.get("font_size_pt", 14), 14)
        max_font_px = max(1, int(round((font_size_pt / 72.0) * px_per_in)))
        alignment = self._alignment_flags()

        for seq_index, spec in enumerate(self.row_specs):
            slot = self.missing + seq_index
            if slot >= MAX_LABELS:
                break
            r = self._label_rect(slot, page_rect)
            painter.setPen(QPen(QColor(120, 124, 130), 1))
            painter.drawRect(r)
            painter.setPen(QPen(QColor(str(self.text_style.get("text_color", "#000000"))), 1))

            text = self._text_from_spec(spec)
            if not text:
                continue

            pad_x = max(1.0, 0.035 * px_per_in)
            pad_y = max(1.0, 0.035 * px_per_in)
            text_x = r.x() + pad_x
            text_y = r.y() + pad_y
            text_w = max(1.0, r.width() - (2.0 * pad_x))
            text_h = max(1.0, r.height() - (2.0 * pad_y))
            font, lines, line_spacing, total_h = self._fit_text_layout(
                text, text_w, text_h, max_font_px
            )
            if not lines:
                continue
            painter.setFont(font)
            line_h = max(1, painter.fontMetrics().height())
            current_y = text_y + max(0.0, (text_h - total_h) / 2.0)

            for line in lines:
                line_rect = QRectF(text_x, current_y, text_w, line_h)
                painter.drawText(line_rect, alignment, line)
                current_y += line_h + line_spacing

        title_rect = QRectF(page_rect.x(), page_rect.y() - 18, page_rect.width(), 14)
        painter.setPen(QPen(QColor(90, 96, 104), 1))
        painter.setFont(QFont("Arial", 8))
        painter.drawText(
            title_rect,
            int(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignVCenter),
            "Avery 5267 text-label preview",
        )

        painter.end()


class Avery5267Window(QMainWindow):
    COL_SLOT = 0
    COL_USE_MARKER = 1
    COL_MARKER_ID = 2
    COL_LINE1 = 3
    COL_LINE2 = 4
    COL_LINE3 = 5
    COL_LINE4 = 6
    COL_LINE5 = 7
    COL_SIDE_LINE = 8
    TEXT_COL_SLOT = 0
    TEXT_COL_TEXT = 1
    # Backward-compatible aliases used in older code paths/saved sessions.
    COL_USE_ARUCO = COL_USE_MARKER
    COL_ARUCO_ID = COL_MARKER_ID
    LINE_COLS = (COL_LINE1, COL_LINE2, COL_LINE3, COL_LINE4, COL_LINE5)
    TEXT_LINE_COLS = (TEXT_COL_TEXT,)
    SETTINGS_FILENAME = ".avery_5267_gui.ini"
    PRINT_DEBUG_FILENAME = ".avery_5267_print_debug.log"

    def __init__(self):
        super(Avery5267Window, self).__init__()
        self._table_syncing = False
        self._text_table_syncing = False
        self._copied_row_payload = None
        self._copied_text_row_payload = None
        self.import_source_path = ""
        self.import_source_kind = ""
        self.import_headers = []
        self.import_rows = []
        self._loading_import_sheet = False
        repo_root = os.path.dirname(SCRIPT_DIR)
        self.settings_path = os.path.join(repo_root, self.SETTINGS_FILENAME)
        self.settings = QSettings(self.settings_path, QSettings.Format.IniFormat)

        self.setWindowTitle("Avery 5267 Label Maker")
        self.resize(1400, 900)

        root = QWidget()
        self.setCentralWidget(root)
        root_layout = QHBoxLayout(root)
        root_layout.setContentsMargins(10, 10, 10, 10)
        self.tabs = QTabWidget()
        root_layout.addWidget(self.tabs)

        self.import_tab = QWidget()
        self.editor_tab = QWidget()
        self.text_labels_tab = QWidget()
        self.tabs.addTab(self.import_tab, "Import")
        self.tabs.addTab(self.editor_tab, "Label Builder")
        self.tabs.addTab(self.text_labels_tab, "Text labels")
        self._build_import_tab()
        self._build_editor_tab()
        self._build_text_labels_tab()

        self._build_menu_bar()

        # Signals
        self.missing_spin.valueChanged.connect(self._on_missing_changed)
        self.count_spin.valueChanged.connect(self._on_count_changed)
        self.refresh_preview_btn.clicked.connect(self.refresh_preview)
        self.apply_defaults_btn.clicked.connect(self.apply_defaults_to_rows)
        self.save_btn.clicked.connect(self.save_files)
        self.print_btn.clicked.connect(self.print_sheet)
        self.table.itemChanged.connect(self._on_table_item_changed)
        self.default_aruco_checkbox.toggled.connect(self._on_default_toggled)
        self.marker_type_combo.currentIndexChanged.connect(self._on_defaults_changed)
        self.aruco_start_spin.valueChanged.connect(self._on_defaults_changed)
        self.name_token_checkbox.toggled.connect(self._on_name_token_changed)
        self.researcher_name_edit.textChanged.connect(self._on_name_token_changed)
        self.preview.wheel_adjust_requested.connect(self._on_preview_wheel_adjust)
        self.text_missing_spin.valueChanged.connect(self._on_text_missing_changed)
        self.text_count_spin.valueChanged.connect(self._on_text_count_changed)
        self.text_refresh_btn.clicked.connect(self.refresh_text_preview)
        self.text_save_btn.clicked.connect(self.save_text_files)
        self.text_print_btn.clicked.connect(self.print_text_sheet)
        self.text_table.itemChanged.connect(self._on_text_table_item_changed)
        self.text_font_size_spin.valueChanged.connect(self._on_text_style_changed)
        self.text_bold_checkbox.toggled.connect(self._on_text_style_changed)
        self.text_italic_checkbox.toggled.connect(self._on_text_style_changed)
        self.text_align_combo.currentIndexChanged.connect(self._on_text_style_changed)
        self.text_color_btn.clicked.connect(self._choose_text_color)
        self.text_preview.wheel_adjust_requested.connect(self._on_text_preview_wheel_adjust)
        self.import_open_btn.clicked.connect(self.import_file_dialog)
        self.import_sheet_combo.currentIndexChanged.connect(self._on_import_sheet_changed)
        self.import_apply_btn.clicked.connect(self.apply_import_to_labels)

        self._load_session_state()
        self._load_text_label_state()
        self._on_missing_changed(self.missing_spin.value())
        self.refresh_preview()
        self.refresh_text_preview()
        self.tabs.setCurrentWidget(self.editor_tab)

    def _print_debug_path(self):
        return os.path.join(os.path.dirname(SCRIPT_DIR), self.PRINT_DEBUG_FILENAME)

    def _write_print_debug(self, message):
        try:
            stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            with open(self._print_debug_path(), "a", encoding="utf-8") as handle:
                handle.write("{0} {1}\n".format(stamp, message))
        except Exception:  # pylint: disable=broad-except
            pass

    def _build_import_tab(self):
        layout = QHBoxLayout(self.import_tab)
        layout.setContentsMargins(8, 8, 8, 8)

        self.import_splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(self.import_splitter)

        left_panel = QWidget()
        left_layout = QVBoxLayout(left_panel)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)
        left_layout.addWidget(QLabel("Imported Data"))

        self.import_table = QTableWidget()
        self.import_table.setAlternatingRowColors(True)
        self.import_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.import_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        self.import_table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.import_table.verticalHeader().setVisible(False)
        left_layout.addWidget(self.import_table, 1)
        self.import_splitter.addWidget(left_panel)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)
        self.import_splitter.addWidget(right_panel)

        self.import_open_btn = QPushButton("Open CSV / Excel")
        right_layout.addWidget(self.import_open_btn)

        self.import_file_label = QLabel("No import file loaded.")
        self.import_file_label.setWordWrap(True)
        right_layout.addWidget(self.import_file_label)

        sheet_form = QFormLayout()
        self.import_sheet_combo = QComboBox()
        self.import_sheet_combo.setEnabled(False)
        sheet_form.addRow("Worksheet", self.import_sheet_combo)
        right_layout.addLayout(sheet_form)

        map_label = QLabel("Column Mapping")
        right_layout.addWidget(map_label)

        map_form = QFormLayout()
        self.import_id_combo = QComboBox()
        self.import_line1_combo = QComboBox()
        self.import_line2_combo = QComboBox()
        self.import_line3_combo = QComboBox()
        self.import_line4_combo = QComboBox()
        self.import_line5_combo = QComboBox()
        self.import_side_combo = QComboBox()
        map_form.addRow("ID", self.import_id_combo)
        map_form.addRow("Line 1", self.import_line1_combo)
        map_form.addRow("Line 2", self.import_line2_combo)
        map_form.addRow("Line 3", self.import_line3_combo)
        map_form.addRow("Line 4", self.import_line4_combo)
        map_form.addRow("Line 5", self.import_line5_combo)
        map_form.addRow("Side Line", self.import_side_combo)
        right_layout.addLayout(map_form)

        self.import_map_combos = {
            "id": self.import_id_combo,
            "line1": self.import_line1_combo,
            "line2": self.import_line2_combo,
            "line3": self.import_line3_combo,
            "line4": self.import_line4_combo,
            "line5": self.import_line5_combo,
            "side_line": self.import_side_combo,
        }
        self._set_import_mapping_headers([])

        self.import_apply_btn = QPushButton("Apply Import To Label Rows")
        right_layout.addWidget(self.import_apply_btn)

        self.import_status_label = QLabel("Load a CSV or Excel file to start.")
        self.import_status_label.setWordWrap(True)
        right_layout.addWidget(self.import_status_label)
        right_layout.addStretch(1)

        self.import_splitter.setStretchFactor(0, 4)
        self.import_splitter.setStretchFactor(1, 2)

    def _build_editor_tab(self):
        layout = QHBoxLayout(self.editor_tab)
        layout.setContentsMargins(8, 8, 8, 8)

        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(self.splitter)

        # Left: preview
        self.preview = SheetPreviewWidget()
        self.splitter.addWidget(self.preview)

        # Right: controls
        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(10)
        self.splitter.addWidget(right_panel)

        self.splitter.setStretchFactor(0, 3)
        self.splitter.setStretchFactor(1, 2)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        right_layout.addLayout(form)

        self.missing_spin = QSpinBox()
        self.missing_spin.setRange(0, MAX_LABELS - 1)
        self.missing_spin.setValue(0)
        form.addRow("Missing Labels", self.missing_spin)

        self.count_spin = QSpinBox()
        self.count_spin.setRange(1, MAX_LABELS)
        self.count_spin.setValue(MAX_LABELS)
        form.addRow("Labels To Print", self.count_spin)

        self.dpi_spin = QSpinBox()
        self.dpi_spin.setRange(300, 2400)
        self.dpi_spin.setSingleStep(100)
        self.dpi_spin.setValue(DEFAULT_DPI)
        form.addRow("DPI", self.dpi_spin)

        self.output_edit = QLineEdit("avery_5267_labels")
        form.addRow("Output Stem", self.output_edit)

        self.token_hint_label = QLabel("Autofill tokens: {date}, {name}, {dict}, {id}")
        form.addRow("", self.token_hint_label)

        name_row = QWidget()
        name_row_layout = QHBoxLayout(name_row)
        name_row_layout.setContentsMargins(0, 0, 0, 0)
        self.name_token_checkbox = QCheckBox("Enable {name}")
        self.name_token_checkbox.setChecked(True)
        self.researcher_name_edit = QLineEdit("")
        self.researcher_name_edit.setPlaceholderText("Researcher name")
        self.researcher_name_edit.setEnabled(self.name_token_checkbox.isChecked())
        name_row_layout.addWidget(self.name_token_checkbox)
        name_row_layout.addWidget(self.researcher_name_edit, 1)
        form.addRow("", name_row)

        self.default_aruco_checkbox = QCheckBox("Use Marker by default")
        self.default_aruco_checkbox.setChecked(False)
        form.addRow("", self.default_aruco_checkbox)

        self.marker_type_combo = QComboBox()
        for label, value in MARKER_TYPE_ITEMS:
            self.marker_type_combo.addItem(label, value)
        form.addRow("Marker Type", self.marker_type_combo)

        self.aruco_start_spin = QSpinBox()
        self.aruco_start_spin.setRange(0, 999999)
        self.aruco_start_spin.setValue(0)
        form.addRow("Marker Start ID", self.aruco_start_spin)

        self.aruco_dict_combo = QComboBox()
        self.aruco_dict_combo.addItems(ARUCO_DICTS)
        default_idx = self.aruco_dict_combo.findText(DEFAULT_ARUCO_DICT)
        if default_idx >= 0:
            self.aruco_dict_combo.setCurrentIndex(default_idx)
        form.addRow("ArUco Dictionary", self.aruco_dict_combo)

        btn_row = QHBoxLayout()
        right_layout.addLayout(btn_row)

        self.apply_defaults_btn = QPushButton("Apply Defaults To Rows")
        btn_row.addWidget(self.apply_defaults_btn)

        self.refresh_preview_btn = QPushButton("Refresh Preview")
        btn_row.addWidget(self.refresh_preview_btn)

        self.save_btn = QPushButton("Save")
        btn_row.addWidget(self.save_btn)

        self.print_btn = QPushButton("Print")
        btn_row.addWidget(self.print_btn)

        table_label = QLabel("Text Entry Table (one row per label)")
        right_layout.addWidget(table_label)

        self.table = QTableWidget()
        self.table.setColumnCount(9)
        self.table.setHorizontalHeaderLabels(
            [
                "Slot",
                "Use Marker",
                "ID",
                "Line 1",
                "Line 2",
                "Line 3",
                "Line 4",
                "Line 5",
                "Side Line",
            ]
        )
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(self.COL_SLOT, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(
            self.COL_USE_ARUCO, QHeaderView.ResizeMode.ResizeToContents
        )
        header.setSectionResizeMode(
            self.COL_ARUCO_ID, QHeaderView.ResizeMode.ResizeToContents
        )
        for col in self.LINE_COLS:
            header.setSectionResizeMode(col, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(self.COL_SIDE_LINE, QHeaderView.ResizeMode.Stretch)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        right_layout.addWidget(self.table, 1)

        self.status_label = QLabel("Ready.")
        right_layout.addWidget(self.status_label)

    def _build_text_labels_tab(self):
        layout = QHBoxLayout(self.text_labels_tab)
        layout.setContentsMargins(8, 8, 8, 8)

        self.text_splitter = QSplitter(Qt.Orientation.Horizontal)
        layout.addWidget(self.text_splitter)

        self.text_preview = TextLabelPreviewWidget()
        self.text_splitter.addWidget(self.text_preview)

        right_panel = QWidget()
        right_layout = QVBoxLayout(right_panel)
        right_layout.setSpacing(10)
        self.text_splitter.addWidget(right_panel)
        self.text_splitter.setStretchFactor(0, 3)
        self.text_splitter.setStretchFactor(1, 2)

        form = QFormLayout()
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        right_layout.addLayout(form)

        self.text_missing_spin = QSpinBox()
        self.text_missing_spin.setRange(0, MAX_LABELS - 1)
        self.text_missing_spin.setValue(0)
        form.addRow("Missing Labels", self.text_missing_spin)

        self.text_count_spin = QSpinBox()
        self.text_count_spin.setRange(1, MAX_LABELS)
        self.text_count_spin.setValue(MAX_LABELS)
        form.addRow("Labels To Print", self.text_count_spin)

        self.text_dpi_spin = QSpinBox()
        self.text_dpi_spin.setRange(300, 2400)
        self.text_dpi_spin.setSingleStep(100)
        self.text_dpi_spin.setValue(DEFAULT_DPI)
        form.addRow("DPI", self.text_dpi_spin)

        self.text_output_edit = QLineEdit("avery_5267_text_labels")
        form.addRow("Output Stem", self.text_output_edit)

        self.text_font_size_spin = QSpinBox()
        self.text_font_size_spin.setRange(2, 72)
        self.text_font_size_spin.setValue(14)
        form.addRow("Max Font Size", self.text_font_size_spin)

        self.text_color_btn = QPushButton("#000000")
        self._set_text_color("#000000")
        form.addRow("Text Color", self.text_color_btn)

        style_row = QWidget()
        style_row_layout = QHBoxLayout(style_row)
        style_row_layout.setContentsMargins(0, 0, 0, 0)
        self.text_bold_checkbox = QCheckBox("Bold")
        self.text_italic_checkbox = QCheckBox("Italic")
        style_row_layout.addWidget(self.text_bold_checkbox)
        style_row_layout.addWidget(self.text_italic_checkbox)
        style_row_layout.addStretch(1)
        form.addRow("Style", style_row)

        self.text_align_combo = QComboBox()
        self.text_align_combo.addItem("Left", "left")
        self.text_align_combo.addItem("Center", "center")
        self.text_align_combo.addItem("Right", "right")
        center_idx = self.text_align_combo.findData("center")
        if center_idx >= 0:
            self.text_align_combo.setCurrentIndex(center_idx)
        form.addRow("Alignment", self.text_align_combo)

        btn_row = QHBoxLayout()
        right_layout.addLayout(btn_row)

        self.text_refresh_btn = QPushButton("Refresh Preview")
        btn_row.addWidget(self.text_refresh_btn)

        self.text_save_btn = QPushButton("Save")
        btn_row.addWidget(self.text_save_btn)

        self.text_print_btn = QPushButton("Print")
        btn_row.addWidget(self.text_print_btn)

        right_layout.addWidget(QLabel("Text Entry Table (one row per label)"))

        self.text_table = QTableWidget()
        self.text_table.setColumnCount(2)
        self.text_table.setHorizontalHeaderLabels(["Slot", "Text"])
        header = self.text_table.horizontalHeader()
        header.setSectionResizeMode(self.TEXT_COL_SLOT, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(self.TEXT_COL_TEXT, QHeaderView.ResizeMode.Stretch)
        self.text_table.verticalHeader().setVisible(False)
        self.text_table.setAlternatingRowColors(True)
        self.text_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.text_table.setSelectionMode(QTableWidget.SelectionMode.ExtendedSelection)
        right_layout.addWidget(self.text_table, 1)

        copy_text_action = QAction("Copy Selected Text Row", self)
        copy_text_action.setShortcut("Ctrl+C")
        copy_text_action.setShortcutContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        copy_text_action.triggered.connect(self.copy_selected_text_row)
        self.text_table.addAction(copy_text_action)

        paste_text_action = QAction("Paste To Selected Text Row", self)
        paste_text_action.setShortcut("Ctrl+V")
        paste_text_action.setShortcutContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        paste_text_action.triggered.connect(self.paste_selected_text_row)
        self.text_table.addAction(paste_text_action)

        self.text_status_label = QLabel("Ready.")
        right_layout.addWidget(self.text_status_label)

    def _set_import_mapping_headers(self, headers):
        cleaned_headers = [str(h).strip() for h in headers if str(h).strip()]
        for combo in self.import_map_combos.values():
            combo.blockSignals(True)
            combo.clear()
            combo.addItem("(None)", None)
            for idx, header_name in enumerate(cleaned_headers):
                combo.addItem(header_name, idx)
            combo.setCurrentIndex(0)
            combo.blockSignals(False)

        self._auto_map_import_columns(cleaned_headers)

    def _auto_map_import_columns(self, headers):
        if not headers:
            return

        mapping_candidates = {
            "id": ("id", "markerid", "arucoid", "datamatrixid", "qrid", "code"),
            "line1": ("line1", "text1", "label1", "name"),
            "line2": ("line2", "text2", "label2"),
            "line3": ("line3", "text3", "label3"),
            "line4": ("line4", "text4", "label4"),
            "line5": ("line5", "text5", "label5", "date"),
            "side_line": ("sideline", "side", "side_text", "edge"),
        }

        normalized = [re.sub(r"[^a-z0-9]+", "", h.lower()) for h in headers]
        used = set()
        for target_key, keys in mapping_candidates.items():
            chosen = None
            for idx, norm in enumerate(normalized):
                if idx in used:
                    continue
                if norm in keys:
                    chosen = idx
                    break
            if chosen is not None:
                combo = self.import_map_combos.get(target_key)
                if combo is not None:
                    combo.setCurrentIndex(chosen + 1)
                used.add(chosen)

    def import_file_dialog(self):
        default_path = str(self.settings.value("import_last_path", "")).strip()
        if default_path and os.path.isdir(default_path):
            start_path = default_path
        elif default_path:
            start_path = os.path.dirname(default_path)
        else:
            start_path = os.path.dirname(SCRIPT_DIR)

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import CSV or Excel",
            start_path,
            "Data Files (*.csv *.xlsx *.xlsm *.xltx *.xltm);;CSV Files (*.csv);;Excel Files (*.xlsx *.xlsm *.xltx *.xltm);;All Files (*)",
        )
        if not file_path:
            self.import_status_label.setText("Import cancelled.")
            return

        try:
            self._load_import_file(file_path)
            self.settings.setValue("import_last_path", file_path)
            self.settings.sync()
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Import Failed", str(exc))
            self.import_status_label.setText("Import failed.")

    def _load_import_file(self, file_path):
        path = os.path.abspath(file_path)
        ext = os.path.splitext(path)[1].lower()
        self.import_source_path = path
        self.import_file_label.setText(path)

        if ext == ".csv":
            headers, rows = self._read_csv_table(path)
            self.import_source_kind = "csv"
            self.import_sheet_combo.blockSignals(True)
            self.import_sheet_combo.clear()
            self.import_sheet_combo.addItem("(CSV)", "")
            self.import_sheet_combo.setCurrentIndex(0)
            self.import_sheet_combo.setEnabled(False)
            self.import_sheet_combo.blockSignals(False)
            self._set_import_data(headers, rows)
            return

        if ext not in (".xlsx", ".xlsm", ".xltx", ".xltm"):
            raise ValueError("Unsupported file type: {0}".format(ext or "(none)"))
        if load_workbook is None:
            raise RuntimeError("Excel import requires openpyxl. Install with: pip install openpyxl")

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet_names = [str(name) for name in workbook.sheetnames]
        finally:
            workbook.close()

        if not sheet_names:
            raise ValueError("No worksheets found in Excel file.")

        self.import_source_kind = "excel"
        self.import_sheet_combo.blockSignals(True)
        self.import_sheet_combo.clear()
        for sheet_name in sheet_names:
            self.import_sheet_combo.addItem(sheet_name, sheet_name)
        self.import_sheet_combo.setEnabled(True)
        self.import_sheet_combo.setCurrentIndex(0)
        self.import_sheet_combo.blockSignals(False)
        self._load_excel_sheet(sheet_names[0])

    def _on_import_sheet_changed(self, _index):  # pylint: disable=unused-argument
        if self._loading_import_sheet:
            return
        if self.import_source_kind != "excel":
            return
        sheet_name = self.import_sheet_combo.currentData()
        if not sheet_name:
            return
        try:
            self._load_excel_sheet(str(sheet_name))
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Import Failed", str(exc))
            self.import_status_label.setText("Import failed.")

    def _load_excel_sheet(self, sheet_name):
        if not self.import_source_path:
            raise ValueError("No Excel file selected.")
        self._loading_import_sheet = True
        try:
            headers, rows = self._read_excel_table(self.import_source_path, sheet_name)
            self._set_import_data(headers, rows)
        finally:
            self._loading_import_sheet = False

    def _read_csv_table(self, file_path):
        rows = None
        decode_error = None
        for encoding in ("utf-8-sig", "utf-8", "cp1252"):
            try:
                with open(file_path, "r", encoding=encoding, newline="") as handle:
                    reader = csv.reader(handle)
                    rows = [list(row) for row in reader]
                break
            except UnicodeDecodeError as exc:
                decode_error = exc
        if rows is None:
            raise RuntimeError("Unable to decode CSV file: {0}".format(decode_error))
        return self._rows_to_table_data(rows)

    def _read_excel_table(self, file_path, sheet_name):
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError("Worksheet not found: {0}".format(sheet_name))
            sheet = workbook[sheet_name]
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()
        return self._rows_to_table_data(rows)

    def _rows_to_table_data(self, rows):
        if not rows:
            return [], []

        max_cols = max(len(row) for row in rows)
        normalized_rows = []
        for row in rows:
            padded = list(row) + [None] * (max_cols - len(row))
            normalized_rows.append(padded)

        while normalized_rows and all(
            (cell is None) or (str(cell).strip() == "") for cell in normalized_rows[-1]
        ):
            normalized_rows.pop()

        if not normalized_rows:
            return [], []

        raw_headers = normalized_rows[0]
        data_rows = normalized_rows[1:]
        headers = []
        used = set()
        for idx, value in enumerate(raw_headers):
            base = str(value).strip() if value is not None else ""
            if not base:
                base = "Column {0}".format(idx + 1)
            candidate = base
            suffix = 2
            while candidate.lower() in used:
                candidate = "{0} ({1})".format(base, suffix)
                suffix += 1
            used.add(candidate.lower())
            headers.append(candidate)

        return headers, data_rows

    def _set_import_data(self, headers, rows):
        self.import_headers = list(headers)
        self.import_rows = list(rows)

        self.import_table.setRowCount(len(self.import_rows))
        self.import_table.setColumnCount(len(self.import_headers))
        self.import_table.setHorizontalHeaderLabels(self.import_headers)
        for row_idx, row_values in enumerate(self.import_rows):
            for col_idx in range(len(self.import_headers)):
                value = row_values[col_idx] if col_idx < len(row_values) else ""
                text = "" if value is None else str(value)
                self.import_table.setItem(row_idx, col_idx, QTableWidgetItem(text))
        self.import_table.resizeColumnsToContents()
        self._set_import_mapping_headers(self.import_headers)
        self.import_status_label.setText(
            "Loaded {0} rows from import (excluding header).".format(len(self.import_rows))
        )

    def _get_import_mapped_value(self, row_values, combo):
        col_idx = combo.currentData()
        if col_idx is None:
            return ""
        try:
            value = row_values[int(col_idx)]
        except (TypeError, ValueError, IndexError):
            return ""
        return "" if value is None else str(value).strip()

    def apply_import_to_labels(self):
        if not self.import_rows or not self.import_headers:
            self.import_status_label.setText("Load data before applying import.")
            return

        max_count = MAX_LABELS - self.missing_spin.value()
        if max_count <= 0:
            self.import_status_label.setText("No printable slots available from current missing count.")
            return

        import_count = min(len(self.import_rows), max_count)
        self.count_spin.setValue(import_count)
        self._sync_table_rows()

        self._table_syncing = True
        try:
            for row in range(import_count):
                data_row = self.import_rows[row]

                id_text = self._get_import_mapped_value(data_row, self.import_id_combo)
                if id_text:
                    id_item = self.table.item(row, self.COL_ARUCO_ID)
                    if id_item is not None:
                        id_item.setText(id_text)

                mapped_lines = [
                    self._get_import_mapped_value(data_row, self.import_line1_combo),
                    self._get_import_mapped_value(data_row, self.import_line2_combo),
                    self._get_import_mapped_value(data_row, self.import_line3_combo),
                    self._get_import_mapped_value(data_row, self.import_line4_combo),
                    self._get_import_mapped_value(data_row, self.import_line5_combo),
                ]
                for idx, col in enumerate(self.LINE_COLS):
                    item = self.table.item(row, col)
                    if item is not None:
                        item.setText(mapped_lines[idx])

                side_value = self._get_import_mapped_value(data_row, self.import_side_combo)
                side_item = self.table.item(row, self.COL_SIDE_LINE)
                if side_item is not None:
                    side_item.setText(side_value)
        finally:
            self._table_syncing = False

        self.refresh_preview()
        self.tabs.setCurrentWidget(self.editor_tab)
        applied = import_count
        self.import_status_label.setText("Applied {0} imported row(s) to labels.".format(applied))
        self.status_label.setText("Imported {0} row(s) into label table.".format(applied))

    def _build_menu_bar(self):
        file_menu = self.menuBar().addMenu("&File")

        import_action = QAction("Import CSV/Excel...", self)
        import_action.setShortcut("Ctrl+I")
        import_action.triggered.connect(self.import_file_dialog)
        file_menu.addAction(import_action)

        file_menu.addSeparator()

        save_labels_action = QAction("Save Labels", self)
        save_labels_action.setShortcut("Ctrl+S")
        save_labels_action.triggered.connect(self.save_files)
        file_menu.addAction(save_labels_action)

        print_action = QAction("Print...", self)
        print_action.setShortcut("Ctrl+P")
        print_action.triggered.connect(self.print_sheet)
        file_menu.addAction(print_action)

        file_menu.addSeparator()

        save_session_action = QAction("Save Session...", self)
        save_session_action.setShortcut("Ctrl+Shift+S")
        save_session_action.triggered.connect(self.save_session_as)
        file_menu.addAction(save_session_action)

        load_session_action = QAction("Load Session...", self)
        load_session_action.setShortcut("Ctrl+O")
        load_session_action.triggered.connect(self.load_session_from_file)
        file_menu.addAction(load_session_action)

        file_menu.addSeparator()

        exit_action = QAction("Exit", self)
        exit_action.setShortcut("Ctrl+Q")
        exit_action.triggered.connect(self.close)
        file_menu.addAction(exit_action)

        edit_menu = self.menuBar().addMenu("&Edit")

        copy_row_action = QAction("Copy Selected Row", self)
        copy_row_action.setShortcut("Ctrl+C")
        copy_row_action.setShortcutContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        copy_row_action.triggered.connect(self.copy_selected_row)
        self.table.addAction(copy_row_action)
        edit_menu.addAction(copy_row_action)

        paste_row_action = QAction("Paste To Selected Row", self)
        paste_row_action.setShortcut("Ctrl+V")
        paste_row_action.setShortcutContext(Qt.ShortcutContext.WidgetWithChildrenShortcut)
        paste_row_action.triggered.connect(self.paste_selected_row)
        self.table.addAction(paste_row_action)
        edit_menu.addAction(paste_row_action)

        view_menu = self.menuBar().addMenu("&View")

        zoom_in_action = QAction("Zoom In", self)
        zoom_in_action.setShortcut("Ctrl+=")
        zoom_in_action.triggered.connect(lambda checked=False: self._zoom_preview(1))
        view_menu.addAction(zoom_in_action)

        zoom_out_action = QAction("Zoom Out", self)
        zoom_out_action.setShortcut("Ctrl+-")
        zoom_out_action.triggered.connect(lambda checked=False: self._zoom_preview(-1))
        view_menu.addAction(zoom_out_action)

        zoom_reset_action = QAction("Reset Zoom", self)
        zoom_reset_action.setShortcut("Ctrl+0")
        zoom_reset_action.triggered.connect(lambda checked=False: self._zoom_preview(0))
        view_menu.addAction(zoom_reset_action)

    def _zoom_preview(self, steps):
        if self.tabs.currentWidget() == self.text_labels_tab:
            preview = self.text_preview
            status_label = self.text_status_label
        else:
            preview = self.preview
            status_label = self.status_label
        if int(steps) == 0:
            preview.reset_zoom()
        else:
            preview.zoom_by_steps(int(steps))
        percent = int(round(preview.zoom_factor * 100.0))
        status_label.setText("Preview zoom: {0}%".format(percent))

    def _on_preview_wheel_adjust(self, steps, ctrl_held, shift_held):
        step_value = int(steps)
        if step_value == 0:
            return
        if ctrl_held:
            self._zoom_preview(step_value)
            return
        if shift_held:
            self.count_spin.setValue(self.count_spin.value() + step_value)
            return
        self.missing_spin.setValue(self.missing_spin.value() + step_value)

    def _on_text_preview_wheel_adjust(self, steps, ctrl_held, shift_held):
        step_value = int(steps)
        if step_value == 0:
            return
        if ctrl_held:
            self._zoom_preview(step_value)
            return
        if shift_held:
            self.text_count_spin.setValue(self.text_count_spin.value() + step_value)
            return
        self.text_missing_spin.setValue(self.text_missing_spin.value() + step_value)

    def _selected_table_rows(self):
        selection_model = self.table.selectionModel()
        if selection_model is None:
            return []
        rows = sorted({index.row() for index in selection_model.selectedRows()})
        if rows:
            return rows
        current_row = self.table.currentRow()
        if current_row >= 0:
            return [current_row]
        return []

    def _get_row_payload(self, row):
        return {
            "use_marker": self._is_row_aruco_enabled(row),
            "id_text": self._get_item_text(row, self.COL_ARUCO_ID),
            "lines": [self._get_item_text(row, col) for col in self.LINE_COLS],
            "side_line": self._get_item_text(row, self.COL_SIDE_LINE),
        }

    def _apply_row_payload(self, row, payload):
        use_aruco = self.table.item(row, self.COL_USE_ARUCO)
        if use_aruco is not None:
            use_marker = _safe_bool(payload.get("use_marker", payload.get("has_aruco", False)))
            use_aruco.setCheckState(
                Qt.CheckState.Checked
                if use_marker
                else Qt.CheckState.Unchecked
            )

        aruco_item = self.table.item(row, self.COL_ARUCO_ID)
        if aruco_item is not None:
            aruco_item.setText(str(payload.get("id_text", payload.get("aruco_id_text", ""))))

        lines = payload.get("lines", [])
        if isinstance(lines, str):
            lines = [lines]
        for i, col in enumerate(self.LINE_COLS):
            item = self.table.item(row, col)
            if item is not None:
                item.setText(str(lines[i]) if i < len(lines) else "")

        side_line_item = self.table.item(row, self.COL_SIDE_LINE)
        if side_line_item is not None:
            side_line_item.setText(str(payload.get("side_line", "")))

    def copy_selected_row(self):
        if self.tabs.currentWidget() == self.text_labels_tab:
            self.copy_selected_text_row()
            return

        self._sync_table_rows()
        rows = self._selected_table_rows()
        if not rows:
            self.status_label.setText("Select a row to copy.")
            return

        source_row = rows[0]
        self._copied_row_payload = self._get_row_payload(source_row)
        self.status_label.setText(
            "Copied row {0}. Select another row and press Ctrl+V.".format(source_row + 1)
        )

    def paste_selected_row(self):
        if self.tabs.currentWidget() == self.text_labels_tab:
            self.paste_selected_text_row()
            return

        if not isinstance(self._copied_row_payload, dict):
            self.status_label.setText("Copy a row first (Ctrl+C).")
            return

        self._sync_table_rows()
        rows = self._selected_table_rows()
        if not rows:
            self.status_label.setText("Select a destination row to paste.")
            return

        self._table_syncing = True
        try:
            for row in rows:
                self._apply_row_payload(row, self._copied_row_payload)
        finally:
            self._table_syncing = False

        self.refresh_preview()
        self.status_label.setText("Pasted to row(s): {0}".format(", ".join(str(r + 1) for r in rows)))

    def _set_text_color(self, color_text):
        color = QColor(str(color_text))
        if not color.isValid():
            color = QColor("#000000")
        hex_text = color.name().upper()
        self.text_color_btn.setText(hex_text)
        self.text_color_btn.setStyleSheet(
            "QPushButton {{ background-color: {0}; color: {1}; }}".format(
                hex_text,
                "#FFFFFF" if color.value() < 120 else "#000000",
            )
        )

    def _choose_text_color(self):
        current = QColor(self.text_color_btn.text())
        if not current.isValid():
            current = QColor("#000000")
        color = QColorDialog.getColor(current, self, "Text Color")
        if not color.isValid():
            return
        self._set_text_color(color.name())
        self.refresh_text_preview()
        self.text_status_label.setText("Text color updated.")

    def _text_style(self):
        return {
            "font_size_pt": self.text_font_size_spin.value(),
            "text_color": self.text_color_btn.text().strip() or "#000000",
            "align": self.text_align_combo.currentData() or "center",
            "bold": self.text_bold_checkbox.isChecked(),
            "italic": self.text_italic_checkbox.isChecked(),
        }

    def _on_text_style_changed(self, _value=None):  # pylint: disable=unused-argument
        self.refresh_text_preview()
        self.text_status_label.setText("Text style updated.")

    def _on_text_table_item_changed(self, item):  # pylint: disable=unused-argument
        if self._text_table_syncing:
            return
        self.refresh_text_preview()

    def _on_text_missing_changed(self, value):
        value = int(value)
        if self.missing_spin.value() != value:
            self.missing_spin.blockSignals(True)
            try:
                self.missing_spin.setValue(value)
            finally:
                self.missing_spin.blockSignals(False)

        max_count = MAX_LABELS - int(value)
        self.text_count_spin.setMaximum(max_count)
        if self.text_count_spin.value() > max_count:
            self.text_count_spin.setValue(max_count)
        self._sync_text_table_rows()
        self.refresh_text_preview()

        self.count_spin.setMaximum(max_count)
        if self.count_spin.value() > max_count:
            self.count_spin.setValue(max_count)
        self._sync_table_rows()
        self.refresh_preview()

    def _on_text_count_changed(self, value):  # pylint: disable=unused-argument
        self._sync_text_table_rows()
        self.refresh_text_preview()

    def _sync_text_table_rows(self):
        self._text_table_syncing = True
        try:
            missing = self.text_missing_spin.value()
            count = self.text_count_spin.value()
            self.text_table.setRowCount(count)

            for row in range(count):
                slot = missing + row
                slot_item = self.text_table.item(row, self.TEXT_COL_SLOT)
                if slot_item is None:
                    slot_item = self._make_text_item(str(slot), editable=False, center=True)
                    self.text_table.setItem(row, self.TEXT_COL_SLOT, slot_item)
                else:
                    slot_item.setText(str(slot))

                for col in self.TEXT_LINE_COLS:
                    if self.text_table.item(row, col) is None:
                        self.text_table.setItem(row, col, self._make_text_item("", editable=True))
        finally:
            self._text_table_syncing = False

    def _collect_text_row_specs(self):
        specs = []
        for row in range(self.text_table.rowCount()):
            specs.append({"text": self._get_text_item_text(row, self.TEXT_COL_TEXT)})
        return specs

    def refresh_text_preview(self):
        self._sync_text_table_rows()
        specs = self._collect_text_row_specs()
        self.text_preview.set_text_preview_data(
            self.text_missing_spin.value(),
            specs,
            self._text_style(),
        )
        self.text_status_label.setText("Preview updated.")

    def _get_text_item_text(self, row, col):
        item = self.text_table.item(row, col)
        return item.text().strip() if item else ""

    def _selected_text_table_rows(self):
        selection_model = self.text_table.selectionModel()
        if selection_model is None:
            return []
        rows = sorted({index.row() for index in selection_model.selectedRows()})
        if rows:
            return rows
        current_row = self.text_table.currentRow()
        if current_row >= 0:
            return [current_row]
        return []

    def _get_text_row_payload(self, row):
        return {"text": self._get_text_item_text(row, self.TEXT_COL_TEXT)}

    def _apply_text_row_payload(self, row, payload):
        if not isinstance(payload, dict):
            text = str(payload)
        elif "text" in payload:
            text = str(payload.get("text", ""))
        else:
            lines = payload.get("lines", [])
            if isinstance(lines, str):
                text = lines
            else:
                text = "\n".join(str(line).strip() for line in lines if str(line).strip())
        item = self.text_table.item(row, self.TEXT_COL_TEXT)
        if item is not None:
            item.setText(text)

    def copy_selected_text_row(self):
        self._sync_text_table_rows()
        rows = self._selected_text_table_rows()
        if not rows:
            self.text_status_label.setText("Select a row to copy.")
            return
        source_row = rows[0]
        self._copied_text_row_payload = self._get_text_row_payload(source_row)
        self.text_status_label.setText(
            "Copied row {0}. Select another row and press Ctrl+V.".format(source_row + 1)
        )

    def paste_selected_text_row(self):
        if not isinstance(self._copied_text_row_payload, dict):
            self.text_status_label.setText("Copy a row first (Ctrl+C).")
            return
        self._sync_text_table_rows()
        rows = self._selected_text_table_rows()
        if not rows:
            self.text_status_label.setText("Select a destination row to paste.")
            return

        self._text_table_syncing = True
        try:
            for row in rows:
                self._apply_text_row_payload(row, self._copied_text_row_payload)
        finally:
            self._text_table_syncing = False

        self.refresh_text_preview()
        self.text_status_label.setText(
            "Pasted to row(s): {0}".format(", ".join(str(r + 1) for r in rows))
        )

    def closeEvent(self, event):  # pylint: disable=invalid-name
        self._save_session_state()
        self._save_text_label_state()
        super(Avery5267Window, self).closeEvent(event)

    def _capture_session_payload(self):
        self._sync_table_rows()
        return {
            "version": 1,
            "missing": self.missing_spin.value(),
            "count": self.count_spin.value(),
            "dpi": self.dpi_spin.value(),
            "output_stem": self.output_edit.text().strip(),
            "name_token_enabled": self.name_token_checkbox.isChecked(),
            "researcher_name": self.researcher_name_edit.text(),
            "default_aruco": self.default_aruco_checkbox.isChecked(),
            "marker_type": _normalize_marker_type(self.marker_type_combo.currentData()),
            "aruco_start_id": self.aruco_start_spin.value(),
            "aruco_dict": self.aruco_dict_combo.currentText(),
            "rows": self._collect_row_specs(),
            "splitter_sizes": self.splitter.sizes(),
            "preview_zoom": self.preview.zoom_factor,
        }

    def _apply_session_payload(self, payload):
        if not isinstance(payload, dict):
            return

        rows_payload = payload.get("rows", [])
        if not isinstance(rows_payload, list):
            rows_payload = []

        default_missing = self.missing_spin.value()
        default_count = self.count_spin.value()
        default_row_count = len(rows_payload) if rows_payload else default_count

        missing = _safe_int(payload.get("missing", default_missing), default_missing)
        missing = max(0, min(MAX_LABELS - 1, missing))

        count = _safe_int(payload.get("count", default_row_count), default_row_count)
        count = max(1, min(MAX_LABELS - missing, count))

        dpi = _safe_int(payload.get("dpi", DEFAULT_DPI), DEFAULT_DPI)
        dpi = max(300, min(2400, dpi))

        output_stem = str(payload.get("output_stem", self.output_edit.text().strip())).strip()
        name_token_enabled = _safe_bool(
            payload.get("name_token_enabled", self.name_token_checkbox.isChecked())
        )
        researcher_name = str(
            payload.get("researcher_name", self.researcher_name_edit.text())
        ).strip()
        default_aruco = _safe_bool(
            payload.get("default_aruco", self.default_aruco_checkbox.isChecked())
        )
        marker_type = _normalize_marker_type(
            payload.get("marker_type", self.marker_type_combo.currentData())
        )
        if "marker_type" not in payload and isinstance(rows_payload, list):
            for row_spec in rows_payload:
                if not isinstance(row_spec, dict):
                    continue
                inferred = _normalize_marker_type(
                    row_spec.get("marker_type", MARKER_TYPE_NONE)
                )
                if inferred != MARKER_TYPE_NONE:
                    marker_type = inferred
                    break
        aruco_start = _safe_int(
            payload.get("aruco_start_id", self.aruco_start_spin.value()),
            self.aruco_start_spin.value(),
        )
        aruco_start = max(0, min(999999, aruco_start))
        aruco_dict = str(payload.get("aruco_dict", self.aruco_dict_combo.currentText())).strip()
        preview_zoom = payload.get("preview_zoom", self.preview.zoom_factor)

        widgets_to_block = [
            self.missing_spin,
            self.count_spin,
            self.dpi_spin,
            self.output_edit,
            self.name_token_checkbox,
            self.researcher_name_edit,
            self.default_aruco_checkbox,
            self.marker_type_combo,
            self.aruco_start_spin,
            self.aruco_dict_combo,
        ]
        for w in widgets_to_block:
            w.blockSignals(True)
        try:
            self.missing_spin.setValue(missing)
            self.count_spin.setMaximum(MAX_LABELS - missing)
            self.count_spin.setValue(count)
            self.dpi_spin.setValue(dpi)
            if output_stem:
                self.output_edit.setText(output_stem)
            self.name_token_checkbox.setChecked(name_token_enabled)
            self.researcher_name_edit.setText(researcher_name)
            self.default_aruco_checkbox.setChecked(default_aruco)
            marker_idx = self.marker_type_combo.findData(marker_type)
            if marker_idx >= 0:
                self.marker_type_combo.setCurrentIndex(marker_idx)
            self.aruco_start_spin.setValue(aruco_start)
            dict_idx = self.aruco_dict_combo.findText(aruco_dict)
            if dict_idx >= 0:
                self.aruco_dict_combo.setCurrentIndex(dict_idx)
        finally:
            for w in widgets_to_block:
                w.blockSignals(False)

        self._sync_table_rows()

        if rows_payload:
            self._table_syncing = True
            try:
                for row in range(min(len(rows_payload), self.table.rowCount())):
                    spec = rows_payload[row] if isinstance(rows_payload[row], dict) else {}
                    lines = spec.get("lines", [])
                    if isinstance(lines, str):
                        lines = [lines]
                    lines = [str(x) for x in lines]
                    marker_type = _normalize_marker_type(spec.get("marker_type", MARKER_TYPE_NONE))
                    if marker_type == MARKER_TYPE_NONE and _safe_bool(spec.get("has_aruco", False)):
                        marker_type = MARKER_TYPE_ARUCO

                    use_aruco_item = self.table.item(row, self.COL_USE_ARUCO)
                    if use_aruco_item is not None:
                        use_aruco_item.setCheckState(
                            Qt.CheckState.Checked
                            if marker_type != MARKER_TYPE_NONE
                            else Qt.CheckState.Unchecked
                        )

                    aruco_item = self.table.item(row, self.COL_ARUCO_ID)
                    if aruco_item is not None:
                        marker_id_value = spec.get(
                            "id_text",
                            spec.get(
                                "marker_payload",
                                spec.get("marker_id", spec.get("aruco_id", row)),
                            ),
                        )
                        aruco_item.setText(_marker_value_text(marker_id_value, row))

                    for i, col in enumerate(self.LINE_COLS):
                        item = self.table.item(row, col)
                        if item is not None:
                            item.setText(lines[i] if i < len(lines) else "")

                    side_line_item = self.table.item(row, self.COL_SIDE_LINE)
                    if side_line_item is not None:
                        side_line_item.setText(str(spec.get("side_line", "")))
            finally:
                self._table_syncing = False

        splitter_sizes = payload.get("splitter_sizes")
        if isinstance(splitter_sizes, str):
            splitter_sizes = [x.strip() for x in splitter_sizes.split(",") if x.strip()]
        if isinstance(splitter_sizes, list) and len(splitter_sizes) >= 2:
            try:
                self.splitter.setSizes([int(splitter_sizes[0]), int(splitter_sizes[1])])
            except (TypeError, ValueError):
                pass

        self.preview.set_zoom_factor(_safe_float(preview_zoom, self.preview.zoom_factor))
        self.refresh_preview()

    def _save_session_state(self):
        payload = self._capture_session_payload()
        self.settings.setValue("window_geometry", self.saveGeometry())
        self.settings.setValue("splitter_sizes", payload.get("splitter_sizes", []))
        self.settings.setValue("missing", payload.get("missing", 0))
        self.settings.setValue("count", payload.get("count", 1))
        self.settings.setValue("dpi", payload.get("dpi", DEFAULT_DPI))
        self.settings.setValue("output_stem", payload.get("output_stem", ""))
        self.settings.setValue("name_token_enabled", payload.get("name_token_enabled", False))
        self.settings.setValue("researcher_name", payload.get("researcher_name", ""))
        self.settings.setValue("default_aruco", payload.get("default_aruco", False))
        self.settings.setValue("marker_type", payload.get("marker_type", MARKER_TYPE_ARUCO))
        self.settings.setValue("aruco_start_id", payload.get("aruco_start_id", 0))
        self.settings.setValue("aruco_dict", payload.get("aruco_dict", DEFAULT_ARUCO_DICT))
        self.settings.setValue("preview_zoom", payload.get("preview_zoom", 1.0))
        self.settings.setValue("rows_json", json.dumps(payload.get("rows", [])))
        self.settings.setValue("session_json", json.dumps(payload))
        self.settings.sync()
        if self.settings.status() != QSettings.Status.NoError:
            raise RuntimeError("Unable to persist settings at: {0}".format(self.settings_path))

    def _load_session_state(self):
        payload = None
        session_json = self.settings.value("session_json", "")
        if session_json:
            try:
                payload = json.loads(str(session_json))
            except Exception:  # pylint: disable=broad-except
                payload = None

        if not isinstance(payload, dict):
            rows_json = self.settings.value("rows_json", "")
            try:
                saved_rows = json.loads(rows_json) if rows_json else []
            except Exception:  # pylint: disable=broad-except
                saved_rows = []
            payload = {
                "missing": self.settings.value("missing", self.missing_spin.value()),
                "count": self.settings.value("count", self.count_spin.value()),
                "dpi": self.settings.value("dpi", DEFAULT_DPI),
                "output_stem": self.settings.value("output_stem", self.output_edit.text()),
                "name_token_enabled": self.settings.value("name_token_enabled", False),
                "researcher_name": self.settings.value("researcher_name", ""),
                "default_aruco": self.settings.value(
                    "default_aruco", self.default_aruco_checkbox.isChecked()
                ),
                "marker_type": self.settings.value(
                    "marker_type", self.marker_type_combo.currentData()
                ),
                "aruco_start_id": self.settings.value(
                    "aruco_start_id", self.aruco_start_spin.value()
                ),
                "aruco_dict": self.settings.value(
                    "aruco_dict", self.aruco_dict_combo.currentText()
                ),
                "preview_zoom": self.settings.value("preview_zoom", 1.0),
                "rows": saved_rows if isinstance(saved_rows, list) else [],
                "splitter_sizes": self.settings.value("splitter_sizes", []),
            }

        self._apply_session_payload(payload)

        geom = self.settings.value("window_geometry")
        if geom is not None:
            try:
                self.restoreGeometry(geom)
            except Exception:  # pylint: disable=broad-except
                pass

        splitter_sizes = self.settings.value("splitter_sizes")
        if isinstance(splitter_sizes, str):
            splitter_sizes = [x.strip() for x in splitter_sizes.split(",") if x.strip()]
        if isinstance(splitter_sizes, list) and len(splitter_sizes) >= 2:
            try:
                self.splitter.setSizes([int(splitter_sizes[0]), int(splitter_sizes[1])])
            except (TypeError, ValueError):
                pass

    def _capture_text_label_payload(self):
        self._sync_text_table_rows()
        return {
            "version": 1,
            "missing": self.missing_spin.value(),
            "count": self.text_count_spin.value(),
            "dpi": self.text_dpi_spin.value(),
            "output_stem": self.text_output_edit.text().strip(),
            "style": self._text_style(),
            "rows": self._collect_text_row_specs(),
            "splitter_sizes": self.text_splitter.sizes(),
            "preview_zoom": self.text_preview.zoom_factor,
        }

    def _apply_text_label_payload(self, payload):
        if not isinstance(payload, dict):
            return

        rows_payload = payload.get("rows", [])
        if not isinstance(rows_payload, list):
            rows_payload = []

        default_count = self.text_count_spin.value()
        default_row_count = len(rows_payload) if rows_payload else default_count
        missing = self.missing_spin.value()
        missing = max(0, min(MAX_LABELS - 1, missing))
        count = _safe_int(payload.get("count", default_row_count), default_row_count)
        count = max(1, min(MAX_LABELS - missing, count))
        dpi = _safe_int(payload.get("dpi", DEFAULT_DPI), DEFAULT_DPI)
        dpi = max(300, min(2400, dpi))
        output_stem = str(payload.get("output_stem", self.text_output_edit.text())).strip()
        style = payload.get("style", {})
        if not isinstance(style, dict):
            style = {}

        widgets_to_block = [
            self.text_missing_spin,
            self.text_count_spin,
            self.text_dpi_spin,
            self.text_output_edit,
            self.text_font_size_spin,
            self.text_bold_checkbox,
            self.text_italic_checkbox,
            self.text_align_combo,
        ]
        for w in widgets_to_block:
            w.blockSignals(True)
        try:
            self.text_missing_spin.setValue(missing)
            self.text_count_spin.setMaximum(MAX_LABELS - missing)
            self.text_count_spin.setValue(count)
            self.text_dpi_spin.setValue(dpi)
            if output_stem:
                self.text_output_edit.setText(output_stem)
            self.text_font_size_spin.setValue(
                max(2, min(72, _safe_int(style.get("font_size_pt", 14), 14)))
            )
            self._set_text_color(style.get("text_color", "#000000"))
            self.text_bold_checkbox.setChecked(_safe_bool(style.get("bold", False)))
            self.text_italic_checkbox.setChecked(_safe_bool(style.get("italic", False)))
            align_idx = self.text_align_combo.findData(str(style.get("align", "center")))
            if align_idx >= 0:
                self.text_align_combo.setCurrentIndex(align_idx)
        finally:
            for w in widgets_to_block:
                w.blockSignals(False)

        self._sync_text_table_rows()
        if rows_payload:
            self._text_table_syncing = True
            try:
                for row in range(min(len(rows_payload), self.text_table.rowCount())):
                    spec = rows_payload[row] if isinstance(rows_payload[row], dict) else {}
                    self._apply_text_row_payload(row, spec)
            finally:
                self._text_table_syncing = False

        splitter_sizes = payload.get("splitter_sizes")
        if isinstance(splitter_sizes, str):
            splitter_sizes = [x.strip() for x in splitter_sizes.split(",") if x.strip()]
        if isinstance(splitter_sizes, list) and len(splitter_sizes) >= 2:
            try:
                self.text_splitter.setSizes([int(splitter_sizes[0]), int(splitter_sizes[1])])
            except (TypeError, ValueError):
                pass
        self.text_preview.set_zoom_factor(
            _safe_float(payload.get("preview_zoom", self.text_preview.zoom_factor), 1.0)
        )
        self.refresh_text_preview()

    def _save_text_label_state(self):
        payload = self._capture_text_label_payload()
        self.settings.setValue("text_session_json", json.dumps(payload))
        self.settings.sync()

    def _load_text_label_state(self):
        session_json = self.settings.value("text_session_json", "")
        if not session_json:
            return
        try:
            payload = json.loads(str(session_json))
        except Exception:  # pylint: disable=broad-except
            return
        self._apply_text_label_payload(payload)

    def save_session_as(self):
        default_path = str(self.settings.value("session_last_path", "")).strip()
        if not default_path:
            default_path = os.path.join(os.path.dirname(SCRIPT_DIR), "label_session.json")

        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Session",
            default_path,
            "Label Session (*.json);;All Files (*)",
        )
        if not file_path:
            self.status_label.setText("Save session cancelled.")
            return
        if not file_path.lower().endswith(".json"):
            file_path += ".json"

        payload = self._capture_session_payload()
        try:
            with open(file_path, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, indent=2, ensure_ascii=False)
            self.settings.setValue("session_last_path", file_path)
            self.settings.sync()
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Save Session Failed", str(exc))
            self.status_label.setText("Save session failed.")
            return

        self.status_label.setText("Session saved: {0}".format(file_path))

    def load_session_from_file(self):
        default_path = str(self.settings.value("session_last_path", "")).strip()
        if not default_path:
            default_path = os.path.join(os.path.dirname(SCRIPT_DIR), "label_session.json")

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Session",
            default_path,
            "Label Session (*.json);;All Files (*)",
        )
        if not file_path:
            self.status_label.setText("Load session cancelled.")
            return

        try:
            with open(file_path, "r", encoding="utf-8") as handle:
                payload = json.load(handle)
            self._apply_session_payload(payload)
            self._save_session_state()
            self.settings.setValue("session_last_path", file_path)
            self.settings.sync()
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Load Session Failed", str(exc))
            self.status_label.setText("Load session failed.")
            return

        self.status_label.setText("Session loaded: {0}".format(file_path))

    def _make_text_item(self, text, editable=True, center=False):
        item = QTableWidgetItem(text)
        flags = (
            Qt.ItemFlag.ItemIsEnabled
            | Qt.ItemFlag.ItemIsSelectable
            | Qt.ItemFlag.ItemNeverHasChildren
        )
        if editable:
            flags |= Qt.ItemFlag.ItemIsEditable
        item.setFlags(flags)
        if center:
            item.setTextAlignment(
                int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
            )
        return item

    def _make_checkbox_item(self, checked):
        item = QTableWidgetItem("")
        flags = (
            Qt.ItemFlag.ItemIsEnabled
            | Qt.ItemFlag.ItemIsSelectable
            | Qt.ItemFlag.ItemIsUserCheckable
            | Qt.ItemFlag.ItemNeverHasChildren
        )
        item.setFlags(flags)
        state = Qt.CheckState.Checked if checked else Qt.CheckState.Unchecked
        item.setCheckState(state)
        item.setTextAlignment(
            int(Qt.AlignmentFlag.AlignCenter | Qt.AlignmentFlag.AlignVCenter)
        )
        return item

    def _get_item_text(self, row, col):
        item = self.table.item(row, col)
        return item.text().strip() if item else ""

    def _is_row_aruco_enabled(self, row):
        item = self.table.item(row, self.COL_USE_ARUCO)
        if item is None:
            return False
        return item.checkState() == Qt.CheckState.Checked

    def _is_row_marker_enabled(self, row):
        return self._is_row_aruco_enabled(row)

    def _on_missing_changed(self, value):
        value = int(value)
        if self.text_missing_spin.value() != value:
            self.text_missing_spin.blockSignals(True)
            try:
                self.text_missing_spin.setValue(value)
            finally:
                self.text_missing_spin.blockSignals(False)

        max_count = MAX_LABELS - int(value)
        self.count_spin.setMaximum(max_count)
        if self.count_spin.value() > max_count:
            self.count_spin.setValue(max_count)
        self._sync_table_rows()
        self.refresh_preview()

        self.text_count_spin.setMaximum(max_count)
        if self.text_count_spin.value() > max_count:
            self.text_count_spin.setValue(max_count)
        self._sync_text_table_rows()
        self.refresh_text_preview()

    def _on_count_changed(self, value):  # pylint: disable=unused-argument
        self._sync_table_rows()
        self.refresh_preview()

    def _on_default_toggled(self, checked):  # pylint: disable=unused-argument
        self.status_label.setText("Defaults updated. Click 'Apply Defaults To Rows' to copy.")

    def _on_name_token_changed(self, _value):  # pylint: disable=unused-argument
        self.researcher_name_edit.setEnabled(self.name_token_checkbox.isChecked())
        self.refresh_preview()
        self.status_label.setText("Token settings updated.")

    def _on_defaults_changed(self, value):  # pylint: disable=unused-argument
        self.refresh_preview()
        self.status_label.setText("Defaults updated. Click 'Apply Defaults To Rows' to copy.")

    def _on_table_item_changed(self, item):  # pylint: disable=unused-argument
        if self._table_syncing:
            return
        self.refresh_preview()

    def _sync_table_rows(self):
        self._table_syncing = True
        try:
            missing = self.missing_spin.value()
            count = self.count_spin.value()
            existing = self.table.rowCount()
            self.table.setRowCount(count)

            for row in range(count):
                slot = missing + row
                slot_item = self.table.item(row, self.COL_SLOT)
                if slot_item is None:
                    slot_item = self._make_text_item(str(slot), editable=False, center=True)
                    self.table.setItem(row, self.COL_SLOT, slot_item)
                else:
                    slot_item.setText(str(slot))

                use_aruco = self.table.item(row, self.COL_USE_ARUCO)
                if use_aruco is None:
                    use_aruco = self._make_checkbox_item(self.default_aruco_checkbox.isChecked())
                    self.table.setItem(row, self.COL_USE_ARUCO, use_aruco)

                aruco_item = self.table.item(row, self.COL_ARUCO_ID)
                if aruco_item is None:
                    default_id = self.aruco_start_spin.value() + row
                    aruco_item = self._make_text_item(str(default_id), editable=True, center=True)
                    self.table.setItem(row, self.COL_ARUCO_ID, aruco_item)

                for col in self.LINE_COLS:
                    if self.table.item(row, col) is None:
                        self.table.setItem(row, col, self._make_text_item("", editable=True))
                if self.table.item(row, self.COL_SIDE_LINE) is None:
                    self.table.setItem(
                        row, self.COL_SIDE_LINE, self._make_text_item("", editable=True)
                    )

            # If row count increased, ensure new rows inherit defaults.
            if count > existing:
                for row in range(existing, count):
                    use_aruco = self.table.item(row, self.COL_USE_ARUCO)
                    if use_aruco is not None:
                        state = (
                            Qt.CheckState.Checked
                            if self.default_aruco_checkbox.isChecked()
                            else Qt.CheckState.Unchecked
                        )
                        use_aruco.setCheckState(state)
        finally:
            self._table_syncing = False

    def apply_defaults_to_rows(self):
        self._table_syncing = True
        try:
            count = self.table.rowCount()
            start_id = self.aruco_start_spin.value()
            state = (
                Qt.CheckState.Checked
                if self.default_aruco_checkbox.isChecked()
                else Qt.CheckState.Unchecked
            )
            for row in range(count):
                use_aruco = self.table.item(row, self.COL_USE_ARUCO)
                if use_aruco is None:
                    use_aruco = self._make_checkbox_item(state == Qt.CheckState.Checked)
                    self.table.setItem(row, self.COL_USE_ARUCO, use_aruco)
                use_aruco.setCheckState(state)

                aruco_item = self.table.item(row, self.COL_ARUCO_ID)
                if aruco_item is None:
                    aruco_item = self._make_text_item("", editable=True, center=True)
                    self.table.setItem(row, self.COL_ARUCO_ID, aruco_item)
                aruco_item.setText(str(start_id + row))
        finally:
            self._table_syncing = False

        self.refresh_preview()
        self.status_label.setText("Defaults copied into rows.")

    def _resolve_line_tokens(self, text, marker_id=None):
        raw_text = str(text) if text is not None else ""
        name_value = ""
        if self.name_token_checkbox.isChecked():
            name_value = self.researcher_name_edit.text().strip()

        values = {
            "date": _format_today_line5(),
            "name": name_value,
            "dict": self.aruco_dict_combo.currentText().strip(),
            "id": _marker_value_text(marker_id, ""),
        }

        def _replace(match):
            token_name = match.group(1).lower()
            return values.get(token_name, match.group(0))

        return TOKEN_PATTERN.sub(_replace, raw_text)

    def _collect_row_specs(self, resolve_tokens=False):
        specs = []
        selected_marker_type = _normalize_marker_type(self.marker_type_combo.currentData())
        for row in range(self.table.rowCount()):
            marker_default = self.aruco_start_spin.value() + row
            marker_text = self._get_item_text(row, self.COL_ARUCO_ID)
            marker_payload = _marker_value_text(marker_text, marker_default)
            side_line = self._get_item_text(row, self.COL_SIDE_LINE)
            lines = [self._get_item_text(row, col) for col in self.LINE_COLS]
            has_marker = self._is_row_marker_enabled(row)
            marker_type = selected_marker_type if has_marker else MARKER_TYPE_NONE
            aruco_id = _numeric_marker_id_or_zero(marker_payload)
            if marker_type == MARKER_TYPE_ARUCO:
                marker_id = aruco_id
            else:
                marker_id = marker_payload
            if resolve_tokens:
                lines = [self._resolve_line_tokens(line, marker_payload) for line in lines]
                side_line = self._resolve_line_tokens(side_line, marker_payload)

            specs.append(
                {
                    "lines": lines,
                    "side_line": side_line,
                    "marker_type": marker_type,
                    "marker_id": marker_id,
                    "marker_payload": marker_payload,
                    "id_text": marker_payload,
                    "marker_dict": self.aruco_dict_combo.currentText(),
                    # Backward compatibility for older payload consumers.
                    "has_aruco": bool(has_marker and marker_type == MARKER_TYPE_ARUCO),
                    "aruco_id": aruco_id,
                    "aruco_dict": self.aruco_dict_combo.currentText(),
                }
            )
        return specs

    def refresh_preview(self):
        self._sync_table_rows()
        specs = self._collect_row_specs(resolve_tokens=True)
        self.preview.set_preview_data(self.missing_spin.value(), specs)
        self.status_label.setText("Preview updated.")

    def _generate_outputs(self, output_stem, save_pdf=True):
        self._sync_table_rows()
        if not output_stem:
            raise ValueError("Please set an output file stem.")

        output_dir = os.path.dirname(output_stem)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        row_specs = self._collect_row_specs(resolve_tokens=True)
        self._write_print_debug(
            "generate_outputs start output={0!r} save_pdf={1} missing={2} count={3} dpi={4} rows={5}".format(
                output_stem,
                save_pdf,
                self.missing_spin.value(),
                self.count_spin.value(),
                self.dpi_spin.value(),
                len(row_specs),
            )
        )

        def label_spec_fn(seq_index, slot_index):  # pylint: disable=unused-argument
            if seq_index < len(row_specs):
                return row_specs[seq_index]
            return {
                "lines": [],
                "side_line": "",
                "marker_type": MARKER_TYPE_NONE,
                "marker_id": 0,
                "marker_dict": self.aruco_dict_combo.currentText(),
                "has_aruco": False,
                "aruco_id": 0,
                "aruco_dict": self.aruco_dict_combo.currentText(),
            }

        outputs = generate_avery_5267_sheet(
            label_spec_fn=label_spec_fn,
            output_stem=output_stem,
            missing=self.missing_spin.value(),
            count=self.count_spin.value(),
            dpi=self.dpi_spin.value(),
            save_pdf=save_pdf,
        )
        png_size = os.path.getsize(outputs["png"]) if os.path.exists(outputs["png"]) else 0
        pdf_path = outputs.get("pdf")
        pdf_size = os.path.getsize(pdf_path) if pdf_path and os.path.exists(pdf_path) else 0
        self._write_print_debug(
            "generate_outputs done png={0!r} png_bytes={1} pdf={2!r} pdf_bytes={3}".format(
                outputs["png"], png_size, pdf_path, pdf_size
            )
        )
        return outputs

    def _generate_text_outputs(self, output_stem, save_pdf=True):
        self._sync_text_table_rows()
        if not output_stem:
            raise ValueError("Please set an output file stem.")

        output_dir = os.path.dirname(output_stem)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)

        row_specs = self._collect_text_row_specs()

        def label_text_fn(seq_index, slot_index):  # pylint: disable=unused-argument
            if seq_index < len(row_specs):
                return row_specs[seq_index].get("text", "")
            return []

        return generate_avery_5267_text_sheet(
            label_text_fn=label_text_fn,
            output_stem=output_stem,
            missing=self.text_missing_spin.value(),
            count=self.text_count_spin.value(),
            dpi=self.text_dpi_spin.value(),
            style=self._text_style(),
            save_pdf=save_pdf,
        )

    def _print_png(self, png_path, dpi=None):
        self._write_print_debug("_print_png start path={0!r}".format(png_path))
        if hasattr(QImageReader, "setAllocationLimit"):
            # 0 disables the default cap (commonly 256 MB), needed for 1200 DPI sheets.
            QImageReader.setAllocationLimit(0)
            self._write_print_debug("QImageReader allocation limit disabled")

        self._write_print_debug("reading PNG")
        reader = QImageReader(png_path)
        image = reader.read()
        if image.isNull():
            err = reader.errorString() if hasattr(reader, "errorString") else "unknown error"
            self._write_print_debug("PNG read failed: {0}".format(err))
            raise RuntimeError(
                "Unable to load print image: {0} ({1})".format(png_path, err)
            )
        self._write_print_debug(
            "PNG read done width={0} height={1}".format(image.width(), image.height())
        )

        self._write_print_debug("creating QPrinter")
        printer = QPrinter(QPrinter.PrinterMode.HighResolution)
        self._write_print_debug("configuring QPrinter")
        printer.setFullPage(True)
        printer.setPageOrientation(QPageLayout.Orientation.Portrait)
        printer.setPageSize(QPageSize(QPageSize.PageSizeId.Letter))
        printer.setResolution(int(dpi if dpi is not None else self.dpi_spin.value()))

        self._write_print_debug("creating QPrintDialog")
        dialog = QPrintDialog(printer, self)
        dialog.setWindowTitle("Print Avery 5267 Sheet")
        self._write_print_debug("opening QPrintDialog")
        QApplication.processEvents()
        dialog_result = dialog.exec()
        self._write_print_debug("QPrintDialog returned {0}".format(dialog_result))
        if dialog_result != QDialog.DialogCode.Accepted:
            return False

        painter = QPainter()
        self._write_print_debug("painter.begin start")
        if not painter.begin(printer):
            self._write_print_debug("painter.begin failed")
            raise RuntimeError("Unable to start print job.")
        self._write_print_debug("painter.begin done")
        try:
            self._write_print_debug("paintRectPixels start")
            target = printer.pageLayout().paintRectPixels(printer.resolution())
            self._write_print_debug(
                "drawImage start target={0}x{1}".format(target.width(), target.height())
            )
            painter.drawImage(target, image)
            self._write_print_debug("drawImage done")
        finally:
            painter.end()
            self._write_print_debug("painter.end done")
        return True

    def save_files(self):
        if self.tabs.currentWidget() == self.text_labels_tab:
            self.save_text_files()
            return

        output_stem = self.output_edit.text().strip()
        try:
            outputs = self._generate_outputs(output_stem)
            self._save_session_state()
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Save Failed", str(exc))
            self.status_label.setText("Save failed.")
            return

        msg = "Generated:\n{0}\n{1}".format(outputs["png"], outputs["pdf"])
        QMessageBox.information(self, "Success", msg)
        self.status_label.setText("Saved files successfully.")

    def save_text_files(self):
        output_stem = self.text_output_edit.text().strip()
        try:
            outputs = self._generate_text_outputs(output_stem)
            self._save_text_label_state()
        except Exception as exc:  # pylint: disable=broad-except
            QMessageBox.critical(self, "Save Failed", str(exc))
            self.text_status_label.setText("Save failed.")
            return

        msg = "Generated:\n{0}\n{1}".format(outputs["png"], outputs["pdf"])
        QMessageBox.information(self, "Success", msg)
        self.text_status_label.setText("Saved text-label files successfully.")

    def print_sheet(self):
        if self.tabs.currentWidget() == self.text_labels_tab:
            self.print_text_sheet()
            return

        self._write_print_debug("----- print_sheet requested -----")
        try:
            temp_stem = os.path.join(tempfile.gettempdir(), "avery_5267_print_job")
            self._write_print_debug("temp_stem={0!r}".format(temp_stem))
            self.status_label.setText("Preparing print image...")
            QApplication.processEvents()
            outputs = self._generate_outputs(temp_stem, save_pdf=False)
            printed = self._print_png(outputs["png"])
            if printed:
                self.status_label.setText("Print job sent.")
                self._write_print_debug("print_sheet done: sent")
            else:
                self.status_label.setText("Print cancelled.")
                self._write_print_debug("print_sheet done: cancelled")
            self._save_session_state()
        except Exception as exc:  # pylint: disable=broad-except
            self._write_print_debug("print_sheet exception: {0}".format(exc))
            self._write_print_debug(traceback.format_exc())
            QMessageBox.critical(self, "Print Failed", str(exc))
            self.status_label.setText(
                "Print failed. Debug log: {0}".format(self._print_debug_path())
            )

    def print_text_sheet(self):
        self._write_print_debug("----- print_text_sheet requested -----")
        try:
            temp_stem = os.path.join(tempfile.gettempdir(), "avery_5267_text_print_job")
            self._write_print_debug("text temp_stem={0!r}".format(temp_stem))
            self.text_status_label.setText("Preparing print image...")
            QApplication.processEvents()
            outputs = self._generate_text_outputs(temp_stem, save_pdf=False)
            printed = self._print_png(outputs["png"], dpi=self.text_dpi_spin.value())
            if printed:
                self.text_status_label.setText("Print job sent.")
                self._write_print_debug("print_text_sheet done: sent")
            else:
                self.text_status_label.setText("Print cancelled.")
                self._write_print_debug("print_text_sheet done: cancelled")
            self._save_text_label_state()
        except Exception as exc:  # pylint: disable=broad-except
            self._write_print_debug("print_text_sheet exception: {0}".format(exc))
            self._write_print_debug(traceback.format_exc())
            QMessageBox.critical(self, "Print Failed", str(exc))
            self.text_status_label.setText(
                "Print failed. Debug log: {0}".format(self._print_debug_path())
            )


def main():
    app = QApplication(sys.argv)
    win = Avery5267Window()
    win.showMaximized()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
